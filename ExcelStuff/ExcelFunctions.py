'''  This Block loads the test paste workbook (check name), variables below
show rows and columns of random pasting locations. '''

'''
from openpyxl import load_workbook
wb2 = load_workbook("TEST.xlsx")
#print (wb2.sheetnames)

Data = wb2.get_sheet_by_name("Data")
#Individual Grabs
Type = Data.cell(row=2 ,column=2).value
Floor_Area = Data.cell(row=2, column=9).value
HLF = Data.cell(row=2, column=13).value
'''

'''Cycles through all available titles in PASTE workbook and prints them to terminal'''
"""


"""

import openpyxl

# COPY
wb = openpyxl.load_workbook("TESTCOPY.xlsx")  # Add file name
sheet = wb["TESTCOPY"]  # Add Sheet name

# PASTE
wb2 = openpyxl.load_workbook("TESTPASTE.xlsx")  # Add file name
sheet1 = wb2["Data"]  # Add Sheet name

'''Going to need variables for each parameter in copyrange, arranged by title cycler above?'''

'''
#startcol = 3
#startRow = 2
#endCol = 3
#end_Row = 10

#9th column, 4th row to paste.
'''
def copyRange(startCol, startRow, endCol, endRow, sheet):
	rangeSelected = []
	for i in range(startRow, endRow + 1, 1):
		rowSelected = []
		for j in range(startCol, endCol + 1, 1):
			result = splitBuildingAndMeasurement(sheet.cell(row=i, column=j).value)
			rowSelected.append(result[0])
			rowSelected.append(result[1])
			rowSelected.append(sheet.cell(row=i, column = 3).value)

			#print(sheet.cell(row = i,column = j).value)
			#if sheet.cell(row = i, column = j).value == "None":

		rangeSelected.append(rowSelected)
	print(rangeSelected)
	return rangeSelected

#see comment on copyrange about paste range parameters.
sheet = copyRange(dictionary)

def pasteRange(startCol, startRow, endCol, endRow, sheet, sheet1):
	rangeSelected = []
	for i in range(startRow, endRow + 1, 1):
		rowSelected = []
		for j in range(startCol, endCol + 1, 1):
			rowSelected.append(sheet1.cell(row=i, column=j).value)
			print(sheet1.cell(row=i, column=j).value)
		# if sheet.cell(row = i, column = j).value == "None":

		rangeSelected.append(rowSelected)
		sheet1.append(rangeSelected)


	#Paste = sheet1.cell(row = i, column = j)
	#print(Paste)
	return rangeSelected


#Building and measurement are stored in the same cell.
#This function separate the two values into their own variables

def splitBuildingAndMeasurement(inputValue):
	words = inputValue.split()
	buildingName = words[0] + " " +  words[1]
	measurement = [2] + " " + words[3] + " " + words[4]

	return [buildingName, measurement]



def is_number(s):
    if s.isdigit():
    	return False
    else:
    	return True


#splitBuildingAndMeasurement("Unit 002 884.27 sq m")
copyRange(1, 2, 1, 57, sheet)

cells = copyRange(12, 2, 12, 57, sheet)

dictionary = {}

for cell in cells:
	result = splitBuildingAndMeasurement(cell[0])
	dictionary[result[0]] = result[1]
print(dictionary)
'''

#For loop that loops through ALL titles

def titles(rangemin, rangemax):
	for Titles in range(rangemin, rangemax):
		Titles_list = {sheet1.cell(row=2, column=Titles).value}
		print(Titles_list)

#print(Type + ("|"), Floor_Area +("|"), HLF + ("|"))

result = copyRange(12,2,12,57,sheet)
'''
copyRange(12,2,12,57,sheet)
pasteRange(9,4,9,57, sheet, sheet1)
#titles(9,28)
def unique(result):
	is_unique = True
	unique_buildings = []
	for results in result:
		is_unique = True
		split = splitBuildingAndMeasurement(results[0])
		for building in unique_buildings:
			if split[0] == building:
				is_unique = False
		if is_unique:
			unique_buildings.append(split[0])

	return unique_buildings

#print(len(unique(result)))

